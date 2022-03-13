import nltk
import pandas
import math
import string
import pickle
from datetime import datetime
import openpyxl
from nltk.corpus import stopwords
from nltk.stem.lancaster import LancasterStemmer
from django.conf import settings, os
from wordcloud import WordCloud, STOPWORDS
from nltk.tokenize import word_tokenize
import pickle


file = open('FYP_Naive_Pickle.txt', 'rb')
Pdata = pickle.load(file)
Training_Set = ['Cfreq', 'Ffreq', 'Rfreq', 'Qfreq', 'Afreq']
Total_U_Words = Pdata['Uwords']
Query_Stop_words = Pdata['Query Stop words']


Stop_Words = set(stopwords.words('english'))
stemmer = LancasterStemmer()


def Review_Cleaning(test_str):
    Test_List = []
    if(isinstance(test_str, datetime)):
        dateTimeObj = datetime.now()
        test_str = dateTimeObj.strftime("%d-%b-%Y (%H:%M:%S.%f)")
    if(isinstance(test_str, int)):
        return "awesome"

    test_str = test_str.translate(str.maketrans('', '', string.punctuation))
    test_str = nltk.word_tokenize(test_str)
    for words in test_str:
        if not words in Stop_Words:
            Test_List.append(stemmer.stem(words))
    return Test_List


def Query_review_Cleaning(test_str):
    Test_List = []
    if(isinstance(test_str, datetime)):
        dateTimeObj = datetime.now()
        test_str = dateTimeObj.strftime("%d-%b-%Y (%H:%M:%S.%f)")
    if(isinstance(test_str, int)):
        return "awesome"

    test_str = test_str.translate(str.maketrans('', '', string.punctuation))
    test_str = nltk.word_tokenize(test_str)
    print(test_str)
    for words in test_str:
        if not words in Query_Stop_words:
            Test_List.append(stemmer.stem(words))
            print(words)
    return Test_List


def Naive_Bayes(Test_sentence, DataSet):
    P_of_words = 1
    for word in Test_sentence:
        Occrance_of_word = DataSet[word] + 1
        dinominator = len(DataSet) + Total_U_Words
        P_of_words = P_of_words * (Occrance_of_word / dinominator)
    return P_of_words


def Labling_Reviews(Name):
    value = ['Cfreq', 'Ffreq', 'Rfreq', 'Qfreq', 'Afreq']
    File_Stored_loc = settings.MEDIA_ROOT + '\static\Files/'
    work_book = openpyxl.load_workbook(File_Stored_loc + Name)
    Reviews_Sheet = work_book.active
    complain = 0
    recommend = 0
    query = 0
    others = 0
    appreciation = 0

    count = 0
    for i in range(2, Reviews_Sheet.max_row):
        review = Reviews_Sheet.cell(row=i, column=1)
        review_Str = review.value
        for k in range(5):
            value[k] = Naive_Bayes(Review_Cleaning(review_Str), Pdata[Training_Set[k]])

        Label_to_Row = Reviews_Sheet.cell(row=i, column=2)
        if(value[0] > value[1] and value[0] > value[2] and value[0] > value[3] and value[0] > value[4]):
            Label_to_Row.value = "complaint"
            complain += 1
        elif(value[1] > value[0] and value[1] > value[2] and value[1] > value[3] and value[1] > value[4]):
            Label_to_Row.value = "Feedback"
            others += 1
        elif(value[2] > value[0] and value[2] > value[1] and value[2] > value[3] and value[2] > value[4]):
            Label_to_Row.value = "Recommend"
            recommend += 1
        elif(value[3] > value[0] and value[3] > value[1] and value[3] > value[2] and value[3] > value[4]):
            Label_to_Row.value = "Query"
            query += 1

        elif(value[4] > value[0] and value[4] > value[1] and value[4] > value[2] and value[4] > value[3]):
            Label_to_Row.value = "Appreciation"
            appreciation += 1

        else:
            pass
        count = count + 1

    print(count)

    work_book.save(File_Stored_loc + Name)
    Wordcloud(Name)
    return complain, recommend, query, others, appreciation


def Wordcloud(Name):
    Review_words = ""
    words_list = []
    File_Stored_loc = settings.MEDIA_ROOT + '\static\Files/'
    work_book = openpyxl.load_workbook(File_Stored_loc + Name)
    Reviews_Sheet = work_book.active
   # stopwords=set(STOPWORDS)
    Stop_Words = set(stopwords.words('english'))
    for i in range(2, Reviews_Sheet.max_row - 10):
        review = Reviews_Sheet.cell(row=i, column=1)
        review_Str = str(review.value)
        review_Str = review_Str.lower()
        review_Str = review_Str.translate(str.maketrans('', '', string.punctuation))
        words = word_tokenize(review_Str)
        for word in words:
            if word not in Stop_Words:
                Review_words = Review_words + str(word) + ' '
    Word_Cloud = WordCloud(width=800, height=800,
                           background_color='white', min_font_size=5, max_words=100).generate(Review_words)

    Name = Name.replace('.xlsx', '')
    Data_Pickle = {}
    Data_Pickle['WordCloud'] = Word_Cloud
    file = open(File_Stored_loc + Name + "1" + ".txt", 'ab')
    pickle.dump(Data_Pickle, file)
    file.close()
